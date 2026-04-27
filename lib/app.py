# -*- coding: utf-8 -*-
"""
Created on Fri Mar 27 21:35:51 2026
@author: aanand
"""

import sys
import os

current_dir = os.path.dirname(os.path.dirname(__file__))
sys.path.append(os.path.join(current_dir, "lib"))

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Any, List, Optional
import pandas as pd
import numpy as np
from pathlib import Path
import joblib
import random
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent
CSV_PATH = PROJECT_ROOT / "output" / "merged_output.csv"
RUNLENGTH_PATH = PROJECT_ROOT / "output" / "runlength_dummy.csv"
RUNLENGTH_FALLBACK_PATH = PROJECT_ROOT / "runlength_dummy.csv"

# ── CONFIG FILES ─────────────────────────────────────────────────────────────
# SET THIS to your actual config folder (the folder shown in your screenshot):
CONFIG_DIR = Path(r"D:\dcu-ml-api\HGI\HGI-One yr\config")

# All files are .csv  (xlsx also supported if ever migrated)
CONFIG_FILES = {
    "config":           "Model Config",
    "crudetags":        "Crude Tags",
    "desired_hgi":      "Desired HGI",
    "dynamic_tag":      "Dynamic Tag",
    "errorCode":        "Error Codes",
    "features":         "Features",
    "formulaTags":      "Formula Tags",
    "graphics":         "Graphics",
    "lastHgi":          "Last HGI",
    "outputTagMapping": "Output Tag Mapping",
}
# ─────────────────────────────────────────────────────────────────────────────

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _config_path(name: str) -> Path:
    """Return the path for a config file — tries .csv first (primary), then .xlsx."""
    path_csv  = CONFIG_DIR / f"{name}.csv"
    if path_csv.exists():
        return path_csv
    path_xlsx = CONFIG_DIR / f"{name}.xlsx"
    if path_xlsx.exists():
        return path_xlsx
    raise HTTPException(
        status_code=404,
        detail=(
            f"Config file '{name}' not found. "
            f"Looked for '{path_csv}' and '{path_xlsx}'. "
            f"Verify CONFIG_DIR in app.py is set to the correct folder."
        )
    )


def _read_excel(path: Path) -> dict:
    """Read an xlsx/csv file and return {sheets: {sheetName: {columns, rows}}}."""
    suffix = path.suffix.lower()
    result = {"sheets": {}, "active_sheet": ""}

    if suffix == ".csv":
        df = pd.read_csv(path, dtype=str).fillna("")
        result["sheets"]["Sheet1"] = {
            "columns": list(df.columns),
            "rows": df.values.tolist(),
        }
        result["active_sheet"] = "Sheet1"
        return result

    wb = openpyxl.load_workbook(path, data_only=True)
    result["active_sheet"] = wb.active.title

    for ws in wb.worksheets:
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            result["sheets"][ws.title] = {"columns": [], "rows": []}
            continue
        headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows_raw[0])]
        data_rows = [
            [str(c) if c is not None else "" for c in row]
            for row in rows_raw[1:]
        ]
        result["sheets"][ws.title] = {"columns": headers, "rows": data_rows}

    return result


def _write_excel(path: Path, sheet_name: str, columns: list, rows: list):
    """Write updated sheet data back to xlsx (preserves other sheets)."""
    suffix = path.suffix.lower()

    if suffix == ".csv":
        df = pd.DataFrame(rows, columns=columns)
        df.to_csv(path, index=False)
        return

    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    ws.append(columns)
    for row in rows:
        ws.append(row)

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
#  PYDANTIC MODELS
# ═══════════════════════════════════════════════════════════════════════════════

class SavePayload(BaseModel):
    sheet: str
    columns: List[str]
    rows: List[List[Any]]


# ═══════════════════════════════════════════════════════════════════════════════
#  EXISTING ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/")
def home():
    return {"message": "DCU ML API is running"}


@app.get("/run-model")
def run_model():
    try:
        df = pd.read_csv(CSV_PATH)
        df = df.fillna(0)
        df["Time"] = pd.to_datetime(df["Time"], errors="coerce")
        trend_df = df.tail(2016).copy()

        # Runlength payload
        rl_path = RUNLENGTH_PATH if RUNLENGTH_PATH.exists() else RUNLENGTH_FALLBACK_PATH
        runlength_payload = {"latest": None, "trend": []}
        try:
            if rl_path.exists():
                rl = pd.read_csv(rl_path)
                rl = rl.dropna(subset=["Time", "tag", "value"], how="any")
                rl["Time"] = pd.to_datetime(rl["Time"], errors="coerce")
                rl["value"] = pd.to_numeric(rl["value"], errors="coerce")
                if "run_day" in rl.columns:
                    rl["run_day"] = pd.to_numeric(rl["run_day"], errors="coerce")
                rl = rl.dropna(subset=["Time"])

                wide = rl.pivot_table(index="Time", columns="tag", values="value", aggfunc="last").reset_index()

                if "run_day" in rl.columns:
                    run_day_df = rl.groupby("Time", as_index=False)["run_day"].max()
                    wide = wide.merge(run_day_df, on="Time", how="left")

                wide = wide.sort_values("Time")

                def _to_float_or_none(x):
                    try:
                        if pd.isna(x):
                            return None
                        return float(x)
                    except Exception:
                        return None

                trend_list = []
                for _, r in wide.iterrows():
                    item = {
                        "time": str(r["Time"]),
                        "TMT": _to_float_or_none(r.get("TMT")),
                        "TMT_forecast": _to_float_or_none(r.get("TMT_forecast")),
                    }
                    if "run_day" in wide.columns:
                        item["run_day"] = None if pd.isna(r.get("run_day")) else int(r.get("run_day"))
                    trend_list.append(item)

                runlength_payload["trend"] = trend_list
                if trend_list:
                    runlength_payload["latest"] = trend_list[-1]
        except Exception as _e:
            runlength_payload["error"] = str(_e)

        numeric_cols = [
            "39FI123", "HGI_DrumRunTime.HOL", "39TI203", "39PI995A",
            "39TI199", "39PI101", "39DI348", "39FI347",
            "BIL.39.CokeDrum_HGI_Pred.IDMS",
            "BIL.39.CokeDrum_PredHGI_Upper.IDMS",
            "BIL.39.CokeDrum_PredHGI_Lower.IDMS",
            "BIL.39.CokeDrum_HeavyWtPercent.IDMS",
            "BIL.39.CokeDrum_MediumWtPercent.IDMS",
            "BIL.39.CokeDrum_LightWtPercent.IDMS"
        ]

        for col in numeric_cols:
            if col in trend_df.columns:
                trend_df[col] = pd.to_numeric(trend_df[col], errors="coerce").fillna(0)

        latest = trend_df.iloc[-1]

        def safe_float(val):
            try:
                result = float(val)
                return 0.0 if (result != result or result == float('inf') or result == float('-inf')) else result
            except (TypeError, ValueError):
                return 0.0

        return {
            "latest": {
                "timestamp": str(latest["Time"]),
                "hgi": safe_float(latest["BIL.39.CokeDrum_HGI_Pred.IDMS"]),
                "upper": safe_float(latest["BIL.39.CokeDrum_PredHGI_Upper.IDMS"]),
                "lower": safe_float(latest["BIL.39.CokeDrum_PredHGI_Lower.IDMS"]),
                "houronline": safe_float(latest["HGI_DrumRunTime.HOL"]),
                "furnacecharge": safe_float(latest["39FI123"]),
                "freshcharge": safe_float(latest["39FI347"]),
                "inlettemp": safe_float(latest["39TI203"]),
                "inletpress": safe_float(latest["39PI995A"]),
                "outlettemp": safe_float(latest["39TI199"]),
                "outletpress": safe_float(latest["39PI101"]),
                "residapi": safe_float(latest["39DI348"]),
                "cokedrum_qw_cum_mgal": safe_float(latest["BIL.39.Cokedrum_QW_cum_Mgal.IDMS"]),
                "cokedrum_qw_duration": safe_float(latest["BIL.39.Cokedrum_QW_duration_hr.IDMS"]),
                "cokedrum_cokeheight": safe_float(latest["BIL.39.CokeDrum_CokeHeight.IDMS"]),
                "cokedrum_outage_predicted": safe_float(latest["BIL.39.CokeDrum_Current_Out.IDMS"]),
                "cokedrum_hour1": safe_float(latest["BIL.39.CokeDrum_Hour1.IDMS"]),
                "cokedrum_hour1_outage": safe_float(latest["BIL.39.CokeDrum_Proj_Out1.IDMS"]),
                "cokedrum_hour2": safe_float(latest["BIL.39.CokeDrum_Hour2.IDMS"]),
                "cokedrum_hour2_outage": safe_float(latest["BIL.39.CokeDrum_Proj_Out2.IDMS"]),
                "cokedrum_desired_outage": safe_float(latest["BIL.39.CokeDrum_Desired_Out.IDMS"]),
                "cokedrum_desired_hr": safe_float(latest["BIL.39.CokeDrum_Hour_calc.IDMS"]),  
                "actualdp": safe_float(latest["D4_PDI_Actual"]),
                "cleandp": safe_float(latest["D4_PDI_Pred"]),
                                               
                "avgcot": 45,
                "specificfuelconsumption": 1.3,
                "specificenergyconsumption": 1.2,
                "capacityutilization": 94,
                "cokeyield": 45,
                "drum_status": {
                    "drum1": latest.get("BIL.39.HGI_D3908_Status.IDMS", "Unknown"),
                    "drum2": latest.get("BIL.39.HGI_D3909_Status.IDMS", "Unknown"),
                },
                "furnace_status": {
                    "furnace1": "Online",
                    "furnace2": "Offline",
                },
                "flags": {
                    "flag1": int(latest.get("HGIflag1", 0)),
                    "flag2": int(latest.get("HGIflag2", 0)),
                    "flag3": int(latest.get("HGIflag3", 0))
                },
                "yields": {
                    "heavy": safe_float(latest["BIL.39.CokeDrum_HeavyWtPercent.IDMS"]),
                    "medium": safe_float(latest["BIL.39.CokeDrum_MediumWtPercent.IDMS"]),
                    "light": safe_float(latest["BIL.39.CokeDrum_LightWtPercent.IDMS"])
                }
            },
            "trend": [
                {
                    "time": str(row["Time"]),
                    "hgi": safe_float(row["BIL.39.CokeDrum_HGI_Pred.IDMS"]),
                    "upper": safe_float(row["BIL.39.CokeDrum_PredHGI_Upper.IDMS"]),
                    "lower": safe_float(row["BIL.39.CokeDrum_PredHGI_Lower.IDMS"]),
                    "houronline": safe_float(row["HGI_DrumRunTime.HOL"]),
                    "furnacecharge": safe_float(row["39FI123"]),
                    "inlettemp": safe_float(row["39TI203"]),
                    "inletpress": safe_float(row["39PI995A"]),
                    "outlettemp": safe_float(row["39TI199"]),
                    "outletpress": safe_float(row["39PI101"]),
                    "residapi": safe_float(row["39DI348"]),
                    "freshcharge": safe_float(row["39FI347"]),
                    "actualdp": float(row["D4_PDI_Actual"]),
                    "cleandp": float(row["D4_PDI_Pred"]),
                    "foulingdp": float(row["D4_Fouling"]),
                    "cokeheight": float(row["SW.FPC.DC.0292.Outage.D102_Coke_Height_Predicted.IDMS"]),
                    "cycletime": float(row["SW.FPC.DC.0292.Outage.D102_Cyletime_hrs.IDMS"]),
                    "outagepredicted": float(row["SW.FPC.DC.0292.Outage.D102_Outage_Predicted.IDMS"]),
                    "cokedrum_qw_cum_mgal": safe_float(row["BIL.39.Cokedrum_QW_cum_Mgal.IDMS"]),
                    "cokedrum_qw_duration": safe_float(row["BIL.39.Cokedrum_QW_duration_hr.IDMS"]),
                    "cokedrum_cokeheight": safe_float(row["BIL.39.CokeDrum_CokeHeight.IDMS"]),
                    "cokedrum_outage_predicted": safe_float(row["BIL.39.CokeDrum_Current_Out.IDMS"]),
                    "cokedrum_hour1": safe_float(row["BIL.39.CokeDrum_Hour1.IDMS"]),
                    "cokedrum_hour1_outage": safe_float(row["BIL.39.CokeDrum_Proj_Out1.IDMS"]),
                    "cokedrum_hour2": safe_float(row["BIL.39.CokeDrum_Hour2.IDMS"]),
                    "cokedrum_hour2_outage": safe_float(row["BIL.39.CokeDrum_Proj_Out2.IDMS"]),
                    "cokedrum_desired_outage": safe_float(row["BIL.39.CokeDrum_Desired_Out.IDMS"]),
                    "cokedrum_desired_hr": safe_float(row["BIL.39.CokeDrum_Hour_calc.IDMS"]), 
                }
                for _, row in trend_df.iterrows()
            ],
            "runlength": runlength_payload
        }

    except Exception as e:
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/config/list")
def list_configs():
    """Return all available config files with metadata."""
    files = []
    for key, label in CONFIG_FILES.items():
        # CSV is primary format; xlsx is fallback
        path_csv  = CONFIG_DIR / f"{key}.csv"
        path_xlsx = CONFIG_DIR / f"{key}.xlsx"
        if path_csv.exists():
            path = path_csv
            ext = "csv"
        elif path_xlsx.exists():
            path = path_xlsx
            ext = "xlsx"
        else:
            path = None
            ext = "csv"  # expected format
        files.append({
            "key": key,
            "label": label,
            "ext": ext,
            "exists": path is not None,
            "size_kb": round(path.stat().st_size / 1024, 1) if path else 0,
            "modified": path.stat().st_mtime if path else None,
        })
    return {"files": files}


@app.get("/config/{name}")
def get_config(name: str):
    """Return the full contents of a config file as JSON."""
    if name not in CONFIG_FILES:
        raise HTTPException(status_code=400, detail=f"Unknown config: {name}")
    path = _config_path(name)
    data = _read_excel(path)
    data["label"] = CONFIG_FILES[name]
    data["key"] = name
    data["ext"] = path.suffix.lstrip(".")   # "csv" or "xlsx"
    return data


@app.post("/config/{name}/save")
def save_config(name: str, payload: SavePayload):
    """Save updated rows back to the config file."""
    if name not in CONFIG_FILES:
        raise HTTPException(status_code=400, detail=f"Unknown config: {name}")
    path = _config_path(name)
    try:
        _write_excel(path, payload.sheet, payload.columns, payload.rows)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"success": True, "message": f"Saved {len(payload.rows)} rows to {name}.xlsx"}